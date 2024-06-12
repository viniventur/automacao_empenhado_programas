import os
from dotenv import load_dotenv, dotenv_values
import warnings
import pandas as pd
import smtplib
from email.message import  EmailMessage
from email.mime.application import MIMEApplication
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import wget
import ssl
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox
import warnings
warnings.filterwarnings('ignore')


data_atual = datetime.now()
data_atual_f = data_atual.date().strftime('%d-%m-%Y')
data_ontem = data_atual - timedelta(days=1)
data_ontem_f = data_ontem.date().strftime('%d-%m-%Y')

base_arquivo = os.path.join('S:/SOP/003 - GERÊNCIA DE ESTUDOS E PROJEÇÕES/SCRIPTS/Bases empenhado - Layne', f'Base Acompanhamento - atualizado {data_atual_f}.xlsx')


meses_padronizados = {
        '1': 'JAN', '2': 'FEV', '3': 'MAR', '4': 'ABR',
        '5': 'MAI', '6': 'JUN', '7': 'JUL', '8': 'AGO',
        '9': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'
    }

# mensagem de concluido
def popup_concluido():
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    messagebox.showinfo("Atualização do empenhado dos programas", "Atualização do empenhado dos programas concluída!")

# mensagem de NAO concluido
def popup_erro():
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    messagebox.showinfo("Atualização do empenhado dos programas", "Atualização não sucedida!\nVerificar possível erro.")

# formatar PTS    
def formatar_PT(s):
    s = str(s).zfill(17)
    return f"{s[0:2]}. {s[2:5]}. {s[5:9]}. {s[9:13]}"

# Padronizando as naturezas
def formatar_natureza(s):
    s = str(s)
    return f"{s[1:-2]}"


def obter_dados():
    try:
        try:
            df_dot = pd.read_excel(f'http://extrator.sefaz.al.gov.br/DESPESAS/COMPARATIVO-DOTACOES/comparativo_dotacao_despesa_2024_siafe_gerado_em_{data_ontem_f}.xlsx')

        except: 
            df_dot = pd.read_excel(f'http://extrator.sefaz.al.gov.br/DESPESAS/COMPARATIVO-DOTACOES/comparativo_dotacao_despesa_2024_siafe_gerado_em_{data_atual_f}.xlsx')
        return df_dot
    except Exception as e:
        print(' ') 
        print('Erro na obtenção dos dados:')
        print(' ')
        print(e)
        popup_erro()
        return None

def tranformar(df_dot):
    try:

        df_f = df_dot[['DESCRICAO_UO', 'PT', 'PT_DESCRICAO', 'NATUREZA6', 'VALOR_EMPENHADO', 'MES', 'PO']]
        for i in ['DESCRICAO_UO', 'PT', 'PT_DESCRICAO', 'NATUREZA6', 'MES', 'PO']:
            df_f[i] = df_f[i].astype(str)
        # concatenando as colunas para o merge - sem PO
        df_x = df_f
        df_x['concat'] = df_x['DESCRICAO_UO']
        for i in ['PT', 'NATUREZA6']:
            df_x['concat'] = df_x['concat'] + df_x[i]
        df_x['concat_po'] = df_x['concat'] + df_x['PO']


        # base com as informacoes
        info = pd.read_excel('S:/SOP/003 - GERÊNCIA DE ESTUDOS E PROJEÇÕES/SCRIPTS/Bases empenhado - Layne/Bases templates/Base acompanhamento Layne.xlsx', sheet_name='atualizada 04-06')

        # fazendo o merge para cada valor e cada mês

        concat_planta = 'SECRETARIA DE ESTADO DA AGRICULTURA, PECUÁRIA, PESCA E AQUICULTURA20605104335800000333903200644'
        concat_leite = 'SECRETARIA DE ESTADO DA AGRICULTURA, PECUÁRIA, PESCA E AQUICULTURA20605104335800000333903200646'

        pos_seagri = [concat_planta, concat_leite]

        for i in sorted(df_f['MES'].unique()):
            info[i] = ''
            teste = df_x.loc[df_x['MES'] == i]
            for j in info['concat'].values:
                if ((j == pos_seagri[0]) | (j == pos_seagri[1])):
                    for po_seagri in pos_seagri:
                        info.loc[info.concat == po_seagri, i] = teste.loc[teste['concat_po'] == po_seagri]['VALOR_EMPENHADO'].sum()
                else:
                    info.loc[info.concat == j, i] = teste.loc[teste.concat == j]['VALOR_EMPENHADO'].sum()

        # Dropando o concat

        info.drop(columns='concat', inplace=True)

        info['TOTAL'] = info.iloc[ : , 5:].sum(axis=1)

        # Padronizando as colunas e organiznado

        info.rename(columns={'DESCRICAO_UO': 'UO', 'PT': 'PROGRAMA DE TRABALHO',
                            'PT_DESCRICAO': 'PROGRAMA DE TRABALHO DESCRIÇÃO',
                                'NATUREZA6': 'NATUREZA'}, inplace=True)


        # padronizacao
        info.rename(columns=meses_padronizados, inplace=True)
        info['PROGRAMA DE TRABALHO'] = info['PROGRAMA DE TRABALHO'].apply(formatar_PT)
        info['NATUREZA'] = info['NATUREZA'].apply(formatar_natureza)

        # Salvando

        info.to_excel(base_arquivo, index=False)
        return info

    except Exception as e:
        print(' ') 
        print('Erro na transformação dos dados:')
        print(' ')
        print(e)
        popup_erro()
        return None

def formatar_planilha(info):
    try:
        modelo_wb = load_workbook('S:/SOP/003 - GERÊNCIA DE ESTUDOS E PROJEÇÕES/SCRIPTS/Bases empenhado - Layne/Bases templates/Base Acompanhamento - Modelo.xlsx')
        modelo_ws = modelo_wb.active

        novo_wb = Workbook()
        novo_ws = novo_wb.active

        # Transferência de dados e formatação do df
        for r_idx, row in enumerate(dataframe_to_rows(info, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = novo_ws.cell(row=r_idx, column=c_idx, value=value)
                modelo_cell = modelo_ws.cell(row=r_idx, column=c_idx)
                if modelo_cell.has_style:
                    cell.font = copy(modelo_cell.font)
                    cell.border = copy(modelo_cell.border)
                    cell.fill = copy(modelo_cell.fill)
                    cell.number_format = copy(modelo_cell.number_format)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Ajustar largura das colunas
        for col in novo_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = max_length + 2
            novo_ws.column_dimensions[column].width = adjusted_width

        # Definir largura específica para a coluna A e D e centralizar todo o conteúdo
        novo_ws.column_dimensions['A'].width = 40
        novo_ws.column_dimensions['B'].width = 30
        novo_ws.column_dimensions['D'].width = 11
        novo_ws.column_dimensions['E'].width = 60

        # Ajuste específico para as colunas E até J
        for col in ['F', 'G', 'H', 'I', 'J', 'K']:
            novo_ws.column_dimensions[col].width = 20 

        # Mesclar células de A19 a D19
        novo_ws.merge_cells('A20:E20')

        # Formatar célula mesclada A19 a D19 e coluna J conforme a célula modelo
        modelo_cell_A20 = modelo_ws['A20']
        novo_ws.merge_cells('A20:E20')
        merged_cell = novo_ws['A20']
        merged_cell.value = "TOTAL"
        merged_cell.font = Font(bold=True)
        merged_cell.font = copy(modelo_cell_A20.font)
        merged_cell.border = copy(modelo_cell_A20.border)
        merged_cell.fill = copy(modelo_cell_A20.fill)
        merged_cell.alignment = copy(modelo_cell_A20.alignment)

        # Inserir fórmula de soma na célula J18 e aplicar formatação do modelo

        novo_ws['K20'] = f"=SUM(K2:K19)"
        modelo_cell_I1 = modelo_ws['J1']
        modelo_cell_I2 = modelo_ws['J2']
        novo_ws['K1'].font = copy(modelo_cell_I1.font)
        novo_ws['K1'].border = copy(modelo_cell_I1.border)
        novo_ws['K1'].fill = copy(modelo_cell_I1.fill)
        novo_ws['K1'].number_format = copy(modelo_cell_I1.number_format)
        novo_ws['K1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        for c in range(6, 13):
            for i in range(2, 21):
                novo_ws.cell(row=i, column=c).font = copy(modelo_cell_I2.font)
                novo_ws.cell(row=i, column=c).border = copy(modelo_cell_I2.border)
                novo_ws.cell(row=i, column=c).fill = copy(modelo_cell_I2.fill)
                novo_ws.cell(row=i, column=c).number_format = copy(modelo_cell_I2.number_format)
                novo_ws.cell(row=i, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Inserir fórmula de soma nas células E19 a I19
        for col in range(6, 13):
            cell = novo_ws.cell(row=20, column=col)
            cell.value = f"=SUM({cell.column_letter}2:{cell.column_letter}19)"

        novo_wb.save(base_arquivo)
    except Exception as e:
        print(' ') 
        print('Erro na formatação da planilha:')
        print(' ')
        print(e)
        popup_erro()

def enviar_email():
    try:

        context = ssl.create_default_context()
        
        load_dotenv()

        conta_email = os.getenv('email_avisos')
        conta_envio = os.getenv('email_sup')
        senha = os.getenv('s_avisos')

        corpo_email = """
        <p><b>Olá, Layne! Segue a planilha de acompanamento de empenho atualizado na data de hoje.<b></p>
        <p>OBS.: esse email é enviado de forma automática.</p>
        """

        msg = EmailMessage()
        msg['Subject'] = f"Acompanhamento de Empenho dos Programas - Atualizado {data_atual_f}"
        msg['From'] = conta_email
        msg['To'] = conta_envio
        # Adicionar o corpo do email como HTML
        msg.add_alternative(corpo_email, subtype='html')

        # Ler o arquivo Excel e anexá-lo ao email
        with open(base_arquivo, 'rb') as file:
            part = MIMEApplication(file.read(), Name=f"Base Acompanhamento - atualizado {data_atual_f}.xlsx")
            part['Content-Disposition'] = f'attachment; filename="Base Acompanhamento - atualizado {data_atual_f}.xlsx"'
            msg.attach(part)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            smtp.login(conta_email, senha)
            smtp.sendmail(conta_email, conta_envio, msg.as_string())

        print(' ')
        print('Atualização finalizada e email enviado!')
        print(' ')
        popup_concluido()

    except Exception as e:
        print(' ')
        print('Atualização finalizada, porém, O EMAIL NÃO FOI ENVIADO:')
        print(' ')
        print(e)
        popup_erro()


print("ATUALIZANDO EMPENHO DOS PROGRAMAS...")

df_dot = obter_dados()
info = tranformar(df_dot)
formatar_planilha(info)
enviar_email()
